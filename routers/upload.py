"""
routers/upload.py — Upload contextual de arquivos no chat do Nicodemus ADM.

Endpoint: POST /chat/upload  (prefixo "/chat" registrado em main.py)
Accept:   multipart/form-data

Form fields:
  file        (UploadFile, obrigatório)
  session_id  (str, obrigatório)
  message     (str, opcional — mensagem do gestor junto ao arquivo)

Tipos suportados: PDF, JPEG, PNG, WebP, XLSX, CSV, DOCX, TXT.
Limite: 10 MB.

Pipeline:
  1. Valida JWT → CurrentUser
  2. Carrega sessão via SessionService.get_or_resume
  3. Rejeita se session.status == "completed" com 400
  4. Valida tipo MIME (415) e tamanho (413)
  5. Extrai conteúdo do arquivo
  6. Classifica o balde de destino (knowledge_base | parent_document |
     transactional | none) via _classify_document_bucket
  7. Se knowledge_base/parent_document: consulta o endpoint /extract
     correspondente na eleve-api, guarda os bytes originais em staging
     (Redis, TTL 15min) e monta doc_suggestion — persistência real só
     acontece após confirmação via POST /chat/kb/confirm
  8. Monta user_text enriquecido
  9. Carrega contexto Redis via memory.get_context
 10. Monta NicoState e invoca nico_graph
 11. Extrai resposta do assistant
 12. Persiste via SessionService.add_message (user e assistant)
 13. await db.commit()
 14. Atualiza Redis via memory.append_turn
 15. Extrai file_id de tool results
 16. Retorna ChatResponse (com doc_suggestion quando aplicável)
"""
from __future__ import annotations

import base64
import csv
import io
import json
import uuid

import docx
import openpyxl
import structlog
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_openai import ChatOpenAI
from redis.asyncio import Redis
from sqlalchemy.ext.asyncio import AsyncSession

from agent.nico_agent import nico_graph
from agent.state import NicoState
from core import memory
from core.api_client import DjangoAPIClient
from core.auth import CurrentUser, get_current_user
from core.database import get_session
from core.settings import settings
from schemas.kb_schemas import DocSuggestion
from schemas.session_types import ChatResponse
from services.session_service import SessionService

logger = structlog.get_logger(__name__)

router = APIRouter()

_MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

# ── Staging de upload pendente de confirmação (knowledge_base / parent_document) ──
# Os bytes originais ficam aqui — nunca entram no loop do LLM.
_KB_UPLOAD_PREFIX = "nicodemus:kb_upload"
_KB_UPLOAD_TTL = 900  # 15 minutos — mesmo padrão de FILE_STORAGE_TTL

# Categorias reais dos models KnowledgeBaseFile/ParentDocument na eleve-api
# (apps/knowledge_base/models.py — KB_CATEGORY_CHOICES / PARENT_DOC_CATEGORY)
_KB_CATEGORIES = [
    "regimento", "calendario", "pedagogico", "financeiro", "matriculas",
    "comunicados", "cardapio", "politicas", "faq", "outros",
]
_PARENT_DOC_CATEGORIES = [
    "material", "manual", "contrato", "formulario", "autorizacao",
    "comunicado", "financeiro", "calendario", "cardapio", "outros",
]

_SUPPORTED_TYPES: dict[str, str] = {
    "application/pdf":                                                          "pdf",
    "image/jpeg":                                                               "image",
    "image/png":                                                                "image",
    "image/webp":                                                               "image",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":       "xlsx",
    "text/csv":                                                                 "csv",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
    "text/plain":                                                               "txt",
}

_vision_llm = ChatOpenAI(
    model="gpt-4o",
    api_key=settings.openai_api_key,
    temperature=0,
    max_tokens=2048,
)

_classify_llm = ChatOpenAI(
    model="gpt-4o-mini",
    api_key=settings.openai_api_key,
    temperature=0,
)


# ── helpers ────────────────────────────────────────────────────────────────────────────────

def _get_redis() -> Redis:
    return Redis.from_url(settings.redis_url, decode_responses=True)


async def _classify_document_bucket(text: str, filename: str) -> dict:
    """
    Classifica o arquivo enviado no chat em um dos baldes de destino:

    knowledge_base  → conhecimento interno, indexado via RAG (regimento, calendário,
                       políticas, cardápio, comunicados internos...) — nunca enviado
                       ao responsável.
    parent_document → documento que a escola ENVIA aos responsáveis, como está
                       (lista de material, contrato, comunicado, autorização...).
    transactional   → comprovante de pagamento, contrato de matrícula ou boletim —
                       já tem fluxo próprio de leitura (/doc/extract), não deve ser
                       persistido por aqui.
    none            → não se encaixa em nenhum dos anteriores.
    """
    prompt = (
        "Classifique o documento a seguir em UM dos baldes abaixo. Responda APENAS "
        'com um JSON: {"bucket": "...", "category": "...", "reason": "frase curta em PT-BR"}\n\n'
        "Baldes:\n"
        "- knowledge_base: conhecimento interno da escola, usado como contexto para "
        f"responder perguntas (category deve ser um destes: {', '.join(_KB_CATEGORIES)})\n"
        "- parent_document: documento que a escola ENVIA aos responsáveis, sem análise "
        f"(category deve ser um destes: {', '.join(_PARENT_DOC_CATEGORIES)})\n"
        "- transactional: comprovante de pagamento, contrato de matrícula ou boletim "
        "de um aluno específico — já tem fluxo próprio, não deve ser salvo aqui "
        '(category: "")\n'
        '- none: não se encaixa em nenhum dos anteriores (category: "")\n\n'
        f"Nome do arquivo: {filename}\n\n"
        f"Conteúdo (trecho):\n{text[:3000]}\n\n"
        "Responda só o JSON, sem markdown."
    )
    try:
        response = await _classify_llm.ainvoke([HumanMessage(content=prompt)])
        result = json.loads(response.content.strip())
        if result.get("bucket") not in ("knowledge_base", "parent_document", "transactional", "none"):
            return {"bucket": "none", "category": "", "reason": ""}
        return result
    except Exception as exc:
        logger.warning("upload.classify_error", filename=filename, error=str(exc))
        return {"bucket": "none", "category": "", "reason": ""}


def _extract_file_id(messages: list[dict]) -> str | None:
    """Extrai o file_id do primeiro resultado de tool que contenha esse campo."""
    for msg in messages:
        if msg.get("role") == "tool":
            try:
                data = json.loads(msg.get("content", "{}"))
                if isinstance(data, dict) and data.get("file_id"):
                    return str(data["file_id"])
            except (json.JSONDecodeError, AttributeError):
                pass
    return None


async def _extract_pdf_or_image(content: bytes, mime_type: str) -> str:
    """Usa GPT-4o Vision para extrair conteúdo textual de PDF ou imagem."""
    b64 = base64.b64encode(content).decode("utf-8")
    data_url = f"data:{mime_type};base64,{b64}"

    response = await _vision_llm.ainvoke([
        SystemMessage(
            content=(
                "Você é um extrator de texto preciso. Extraia e retorne todo o conteúdo "
                "textual do documento ou imagem fornecido, preservando a estrutura original. "
                "Inclua datas, valores, nomes, tabelas e qualquer informação relevante. "
                "Responda em português, mantendo os dados originais sem inventar nada."
            )
        ),
        HumanMessage(
            content=[
                {
                    "type": "image_url",
                    "image_url": {"url": data_url, "detail": "high"},
                }
            ]
        ),
    ])
    return response.content.strip()


def _extract_xlsx(content: bytes) -> str:
    """Extrai até 3 abas × 50 linhas de uma planilha XLSX."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    parts: list[str] = []
    for sheet_name in wb.sheetnames[:3]:
        ws = wb[sheet_name]
        parts.append(f"=== Aba: {sheet_name} ===")
        rows_extracted = 0
        for row in ws.iter_rows(values_only=True):
            if rows_extracted >= 50:
                break
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                parts.append("\t".join(cells))
                rows_extracted += 1
    return "\n".join(parts)


def _extract_csv(content: bytes) -> str:
    """Extrai até 50 linhas de um arquivo CSV."""
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for i, row in enumerate(reader):
        if i >= 50:
            break
        lines.append(",".join(row))
    return "\n".join(lines)


def _extract_docx(content: bytes) -> str:
    """Extrai texto dos parágrafos de um arquivo DOCX."""
    doc = docx.Document(io.BytesIO(content))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _extract_txt(content: bytes) -> str:
    """Decodifica um arquivo TXT (máx 5 000 chars)."""
    return content.decode("utf-8", errors="replace")[:5000]


async def _extract_content(content: bytes, mime_type: str, file_kind: str) -> str:
    """Despacha a extração de conteúdo pelo tipo de arquivo."""
    if file_kind in ("pdf", "image"):
        return await _extract_pdf_or_image(content, mime_type)
    if file_kind == "xlsx":
        return _extract_xlsx(content)
    if file_kind == "csv":
        return _extract_csv(content)
    if file_kind == "docx":
        return _extract_docx(content)
    if file_kind == "txt":
        return _extract_txt(content)
    raise ValueError(f"Tipo de arquivo não suportado internamente: {file_kind}")


# ── endpoint ────────────────────────────────────────────────────────────────────────────────

@router.post("/upload", response_model=ChatResponse)
async def upload_file(
    file: UploadFile = File(...),
    session_id: str = Form(...),
    message: str = Form(default=""),
    user: CurrentUser = Depends(get_current_user),
    db: AsyncSession = Depends(get_session),
):
    """
    Recebe um arquivo via multipart/form-data, extrai conteúdo e
    aciona o Nicodemus para análise contextual.

    Requer sessão ativa — crie via POST /sessions/.
    Sessões encerradas (status=completed) são rejeitadas com 400.
    """
    # ── Validações do arquivo ──────────────────────────────────────────────────
    mime_type = file.content_type or ""
    file_kind = _SUPPORTED_TYPES.get(mime_type)
    if not file_kind:
        raise HTTPException(
            status_code=415,
            detail=(
                f"Tipo de arquivo não suportado: '{mime_type}'. "
                "Tipos aceitos: PDF, JPEG, PNG, WebP, XLSX, CSV, DOCX, TXT."
            ),
        )

    content = await file.read()
    if len(content) > _MAX_FILE_SIZE:
        size_mb = len(content) / 1024 / 1024
        raise HTTPException(
            status_code=413,
            detail=f"Arquivo muito grande ({size_mb:.1f} MB). O limite é 10 MB.",
        )

    filename = file.filename or "arquivo"

    logger.info(
        "upload.received",
        session_id=session_id,
        user_id=user.user_id,
        filename=filename,
        mime_type=mime_type,
        size_bytes=len(content),
    )

    # ── Sessão ────────────────────────────────────────────────────────────────
    try:
        session = await SessionService.get_or_resume(db, session_id, user.user_id)
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc))

    if session.status == "completed":
        raise HTTPException(
            status_code=400,
            detail="Sessão encerrada. Crie uma nova sessão para continuar.",
        )

    # ── Extração de conteúdo ──────────────────────────────────────────────────
    try:
        extracted = await _extract_content(content, mime_type, file_kind)
    except Exception as exc:
        logger.error("upload.extraction_error", filename=filename, error=str(exc))
        raise HTTPException(
            status_code=422,
            detail=f"Não foi possível extrair o conteúdo do arquivo: {exc}",
        )

    # ── Classificação do balde de destino ────────────────────────────────────────
    classification = await _classify_document_bucket(extracted, filename)
    bucket = classification.get("bucket", "none")
    category = classification.get("category", "")
    reason = classification.get("reason", "")
    msg_text = message.strip()

    doc_suggestion: dict | None = None

    if bucket in ("knowledge_base", "parent_document"):
        extract_path = (
            "/api/v1/knowledge-base/files/extract/"
            if bucket == "knowledge_base"
            else "/api/v1/knowledge-base/parent-docs/extract/"
        )
        extract_result: dict = {}
        try:
            async with DjangoAPIClient(token=user.sa_token) as client:
                extract_result = await client.post_multipart(
                    extract_path,
                    files={"file": (filename, content, mime_type)},
                )
        except Exception as exc:
            logger.error(
                "upload.kb_extract_error", bucket=bucket, filename=filename, error=str(exc)
            )

        upload_id = str(uuid.uuid4())
        suggestion = DocSuggestion(
            upload_id=upload_id,
            bucket=bucket,
            filename=filename,
            suggested_name=extract_result.get("suggested_name") or filename,
            suggested_category=extract_result.get("suggested_category") or category or "outros",
            description=(extract_result.get("description") or extract_result.get("summary") or "")[:300],
            tags=extract_result.get("suggested_tags", []),
            trigger_phrases=extract_result.get("suggested_trigger_phrases", []),
            audiences=extract_result.get("suggested_audiences", []),
            reason=reason,
        )
        doc_suggestion = suggestion.model_dump()

        # Staging dos bytes originais — nunca entram no loop do LLM.
        staged_payload = json.dumps({
            "school_id": user.school_id,
            "bucket": bucket,
            "filename": filename,
            "mime_type": mime_type,
            "file_b64": base64.b64encode(content).decode("utf-8"),
            "suggested": doc_suggestion,
        })
        r = _get_redis()
        try:
            await r.set(f"{_KB_UPLOAD_PREFIX}:{upload_id}", staged_payload, ex=_KB_UPLOAD_TTL)
        finally:
            await r.aclose()

        logger.info(
            "upload.kb_staged",
            upload_id=upload_id,
            bucket=bucket,
            category=doc_suggestion["suggested_category"],
        )

    # ── Prompt enriquecido ────────────────────────────────────────────────────────
    prompt_lines = [
        f"[ARQUIVO ENVIADO: {filename}]",
        "Conteúdo extraído:",
        extracted,
    ]

    if bucket == "knowledge_base":
        prompt_lines.append(
            "\n[SUGESTÃO DA LITA] Este arquivo parece ser conhecimento interno da escola "
            f"(categoria sugerida: {doc_suggestion['suggested_category']}). "
            f"Nome sugerido: \"{doc_suggestion['suggested_name']}\". "
            "Apresente essa sugestão ao gestor e explique que, se ele confirmar, o "
            "arquivo será salvo na base de conhecimento (indexado para busca semântica) "
            "via POST /chat/kb/confirm — nunca é enviado diretamente aos responsáveis."
        )
    elif bucket == "parent_document":
        prompt_lines.append(
            "\n[SUGESTÃO DA LITA] Este arquivo parece ser um documento para enviar aos "
            f"responsáveis (categoria sugerida: {doc_suggestion['suggested_category']}). "
            f"Nome sugerido: \"{doc_suggestion['suggested_name']}\". "
            "Apresente essa sugestão ao gestor e explique que, se ele confirmar, o "
            "documento fica disponível para envio automático aos responsáveis via "
            "POST /chat/kb/confirm."
        )
    elif bucket == "transactional":
        prompt_lines.append(
            "\n[SUGESTÃO DA LITA] Este arquivo parece ser um comprovante de pagamento, "
            "contrato de matrícula ou boletim de um aluno específico. Oriente o gestor "
            "a usar a tela de leitura de documentos (upload → extração → confirmação) "
            "já existente no painel para registrar isso, em vez de tratar por aqui."
        )

    if msg_text:
        prompt_lines.append(f"\nMensagem do gestor: {msg_text}")
    elif bucket == "none":
        prompt_lines.append(
            "\nAnalise o conteúdo acima, explique o que encontrou e sugira "
            "próximas ações possíveis para a gestão escolar."
        )

    user_text = "\n".join(prompt_lines)

    # ── Invocação do agente ───────────────────────────────────────────────────
    context = await memory.get_context(session_id)
    user_msg_for_agent = {"role": "user", "content": user_text}
    messages_for_agent = context + [user_msg_for_agent]

    logger.info(
        "upload.invoke",
        session_id=session_id,
        user_id=user.user_id,
        bucket=bucket,
        category=category,
        msg_count=len(messages_for_agent),
    )

    initial_state: NicoState = {
        "user_id":      user.user_id,
        "school_id":    user.school_id,
        "sa_token":     user.sa_token,
        "role":         user.role,
        "user_name":    user.name,
        "session_id":   session_id,
        "messages":     messages_for_agent,
        "user_message": user_text,
        "tool_calls":   [],
        "response":     "",
        "error":        None,
    }

    final_state: NicoState = await nico_graph.ainvoke(initial_state)

    reply = final_state.get("response") or ""
    if not reply:
        reply = "Não consegui processar o arquivo. Tente novamente."

    # ── Persistência ────────────────────────────────────────────────────────────
    all_messages = final_state.get("messages", [])
    file_id = _extract_file_id(all_messages)
    file_url = f"/report/download/{file_id}" if file_id else None

    # Mensagem do usuário no histórico mostra [Arquivo: nome.pdf]
    display_content = f"[Arquivo: {filename}]"
    if msg_text:
        display_content += f"\n{msg_text}"

    await SessionService.add_message(db, session, "user", display_content)
    await SessionService.add_message(
        db,
        session,
        "assistant",
        reply,
        metadata={"file_id": file_id} if file_id else None,
    )

    await db.commit()

    # Atualiza Redis com o par user/assistant
    redis_user_msg = {"role": "user", "content": display_content}
    assistant_msg = {"role": "assistant", "content": reply}
    await memory.append_turn(session_id, redis_user_msg, assistant_msg)

    if file_id:
        await SessionService.increment_report_count(db, session)

    logger.info(
        "upload.ok",
        session_id=session_id,
        user_id=user.user_id,
        bucket=bucket,
        category=category,
        has_file=file_id is not None,
        has_doc_suggestion=doc_suggestion is not None,
    )

    return ChatResponse(
        session_id=session_id,
        reply=reply,
        file_id=file_id,
        file_url=file_url,
        doc_suggestion=doc_suggestion,
    )
